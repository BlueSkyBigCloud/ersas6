import csv
import io
import os

from openpyxl import load_workbook

from .models import DataImport


# ============================================================
# File type helpers
# ============================================================

def get_file_extension(filename):
    """
    Return the lowercase file extension.
    """

    return os.path.splitext(filename)[1].lower()


def get_importer(source_type):
    """
    Return the appropriate importer based on the DataImport
    source type.
    """

    importers = {
        DataImport.SourceType.CSV: read_csv,
        DataImport.SourceType.XLSX: read_xlsx,
        DataImport.SourceType.XLS: read_xls,
        DataImport.SourceType.DAT: read_dat,
    }

    importer = importers.get(source_type)

    if importer is None:
        raise ValueError(
            f"Unsupported source type: {source_type}"
        )

    return importer


# ============================================================
# File encoding
# ============================================================

def decode_file(file_data):
    """
    Decode uploaded file bytes.

    UTF-8 is attempted first. If that fails, fall back to
    common Windows encodings.
    """

    if isinstance(file_data, str):
        return file_data

    encodings = [
        "utf-8-sig",
        "utf-8",
        "cp1252",
        "latin-1",
    ]

    for encoding in encodings:

        try:
            return file_data.decode(encoding)

        except UnicodeDecodeError:
            continue

    raise ValueError(
        "Unable to decode the uploaded file."
    )


# ============================================================
# Header normalization
# ============================================================

def normalize_header(header):
    """
    Normalize a source column name.

    We preserve the actual header text while removing
    unnecessary whitespace.
    """

    if header is None:
        return ""

    return str(header).strip()


def normalize_row(row):
    """
    Normalize the keys and values in a row.
    """

    normalized = {}

    for key, value in row.items():

        key = normalize_header(key)

        if not key:
            continue

        if isinstance(value, str):
            value = value.strip()

        normalized[key] = value

    return normalized


# ============================================================
# CSV
# ============================================================

def read_csv(file_data):
    """
    Read CSV data and return a list of dictionaries.

    Example:

        [
            {
                "Emp #": "EMP001",
                "First": "John",
                "Last": "Smith",
            },
            ...
        ]
    """

    text = decode_file(file_data)

    stream = io.StringIO(
        text,
        newline="",
    )

    # --------------------------------------------------------
    # Detect delimiter
    # --------------------------------------------------------

    try:

        sample = text[:4096]

        dialect = csv.Sniffer().sniff(
            sample,
            delimiters=",;\t|",
        )

    except csv.Error:

        dialect = csv.excel

    reader = csv.DictReader(
        stream,
        dialect=dialect,
    )

    if not reader.fieldnames:
        raise ValueError(
            "The CSV file does not contain column headers."
        )

    headers = [
        normalize_header(header)
        for header in reader.fieldnames
    ]

    if not any(headers):
        raise ValueError(
            "The CSV file does not contain valid column headers."
        )

    rows = []

    for row in reader:

        normalized_row = normalize_row(row)

        # ----------------------------------------------------
        # Skip completely empty rows
        # ----------------------------------------------------

        if not any(
            value not in ("", None)
            for value in normalized_row.values()
        ):
            continue

        rows.append(
            normalized_row
        )

    return rows


# ============================================================
# XLSX
# ============================================================

def read_xlsx(file_data):
    """
    Read an XLSX workbook.

    The first worksheet is used as the import source.
    The first row is treated as the column header row.
    """

    workbook = load_workbook(
        filename=io.BytesIO(file_data),
        read_only=True,
        data_only=True,
    )

    try:

        worksheet = workbook.active

        rows_iterator = worksheet.iter_rows(
            values_only=True
        )

        try:
            header_row = next(
                rows_iterator
            )

        except StopIteration:

            raise ValueError(
                "The Excel file is empty."
            )

        headers = [
            normalize_header(value)
            for value in header_row
        ]

        # ----------------------------------------------------
        # Remove trailing empty columns
        # ----------------------------------------------------

        while headers and not headers[-1]:
            headers.pop()

        if not headers:

            raise ValueError(
                "The Excel file does not contain column headers."
            )

        rows = []

        for row in rows_iterator:

            row_values = list(row)

            # Make the row the same length as the headers.
            if len(row_values) < len(headers):

                row_values.extend(
                    [None] *
                    (len(headers) - len(row_values))
                )

            row_values = row_values[
                :len(headers)
            ]

            normalized_row = {}

            for index, header in enumerate(headers):

                if not header:
                    continue

                value = row_values[index]

                if isinstance(value, str):
                    value = value.strip()

                normalized_row[header] = value

            # ------------------------------------------------
            # Skip completely empty rows
            # ------------------------------------------------

            if not any(
                value not in ("", None)
                for value in normalized_row.values()
            ):
                continue

            rows.append(
                normalized_row
            )

        return rows

    finally:

        workbook.close()


# ============================================================
# XLS
# ============================================================

def read_xls(file_data):
    """
    Read legacy .xls files.

    xlrd is intentionally imported here rather than globally
    so the application can still run when only XLSX support
    is installed.
    """

    try:
        import xlrd

    except ImportError:

        raise ValueError(
            "Legacy .xls files require the 'xlrd' package."
        )

    workbook = xlrd.open_workbook(
        file_contents=file_data,
    )

    if workbook.nsheets == 0:

        raise ValueError(
            "The Excel file does not contain any worksheets."
        )

    worksheet = workbook.sheet_by_index(0)

    if worksheet.nrows == 0:

        raise ValueError(
            "The Excel file is empty."
        )

    headers = [
        normalize_header(
            worksheet.cell_value(0, column)
        )
        for column in range(
            worksheet.ncols
        )
    ]

    while headers and not headers[-1]:
        headers.pop()

    if not headers:

        raise ValueError(
            "The Excel file does not contain column headers."
        )

    rows = []

    for row_index in range(
        1,
        worksheet.nrows,
    ):

        normalized_row = {}

        for column_index, header in enumerate(
            headers
        ):

            if not header:
                continue

            value = worksheet.cell_value(
                row_index,
                column_index,
            )

            if isinstance(value, str):
                value = value.strip()

            normalized_row[header] = value

        if not any(
            value not in ("", None)
            for value in normalized_row.values()
        ):
            continue

        rows.append(
            normalized_row
        )

    return rows


# ============================================================
# DAT
# ============================================================

def read_dat(file_data):
    """
    Read a .dat file.

    Many .dat exports are actually delimited text files.
    We attempt delimiter detection rather than assuming a
    specific delimiter.
    """

    text = decode_file(file_data)

    stream = io.StringIO(
        text,
        newline="",
    )

    try:

        sample = text[:4096]

        dialect = csv.Sniffer().sniff(
            sample,
            delimiters=",;\t|",
        )

    except csv.Error:

        # Most common fallback for DAT exports.
        dialect = csv.excel_tab

    reader = csv.DictReader(
        stream,
        dialect=dialect,
    )

    if not reader.fieldnames:

        raise ValueError(
            "The DAT file does not contain column headers."
        )

    rows = []

    for row in reader:

        normalized_row = normalize_row(row)

        if not any(
            value not in ("", None)
            for value in normalized_row.values()
        ):
            continue

        rows.append(
            normalized_row
        )

    return rows


# ============================================================
# Main reader
# ============================================================

def read_import_rows(data_import):
    """
    Read a DataImport file and return normalized rows.

    This is the primary function used by validators.py.

    Example return value:

        [
            {
                "Emp #": "EMP001",
                "First": "John",
                "Last": "Smith",
            },
            {
                "Emp #": "EMP002",
                "First": "Jane",
                "Last": "Doe",
            },
        ]
    """

    if not data_import.file:
        raise ValueError(
            "This import does not have an uploaded file."
        )

    importer = get_importer(
        data_import.source_type
    )

    # --------------------------------------------------------
    # Make sure the file is positioned at the beginning.
    # --------------------------------------------------------

    data_import.file.open(
        mode="rb"
    )

    try:

        file_data = data_import.file.read()

    finally:

        data_import.file.close()

    if not file_data:
        raise ValueError(
            "The uploaded file is empty."
        )

    rows = importer(
        file_data
    )

    return rows


# ============================================================
# File analysis
# ============================================================

def analyze_import_file(data_import):
    """
    Analyze an uploaded file without performing validation.

    Returns information used by the analyze view.
    """

    rows = read_import_rows(
        data_import
    )

    if not rows:

        return {
            "headers": [],
            "total_rows": 0,
            "sample_rows": [],
            "column_count": 0,
        }

    headers = list(
        rows[0].keys()
    )

    return {
        "headers": headers,
        "total_rows": len(rows),
        "sample_rows": rows[:10],
        "column_count": len(headers),
    }